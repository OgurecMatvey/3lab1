"""
Источники данных о курсе валют.

Применённые принципы:
- Абстракция: DataSource описывает контракт "load() -> список CurrencyRate"
- Наследование/полиморфизм: CsvDataSource и ExcelDataSource по-разному
  реализуют load(), но взаимозаменяемы для остального кода (frontend
  работает только с типом DataSource).
- Открытость/закрытость (SOLID, O): чтобы добавить новый формат файла,
  достаточно унаследоваться от DataSource, не меняя остальной код.

Автор: backend (участник 1)
"""
from __future__ import annotations

import csv
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import List

from backend.models import CurrencyRate


class DataSource(ABC):
    """Абстрактный источник данных о курсах валют."""

    @abstractmethod
    def load(self, path: str) -> List[CurrencyRate]:
        """Считать файл и вернуть список курсов по дням, отсортированный по дате."""
        raise NotImplementedError

    @staticmethod
    def create_for(path: str) -> "DataSource":
        """Фабричный метод: подобрать источник данных по расширению файла."""
        suffix = Path(path).suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return ExcelDataSource()
        return CsvDataSource()


class CsvDataSource(DataSource):
    """Чтение курсов валют из CSV-файла вида: date,USD,EUR"""

    def load(self, path: str) -> List[CurrencyRate]:
        rows: List[CurrencyRate] = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            currencies = [c for c in reader.fieldnames if c.lower() != "date"]
            for row in reader:
                day = _parse_date(row["date"])
                rates = {c: float(row[c].replace(",", ".")) for c in currencies}
                rows.append(CurrencyRate(day=day, rates=rates))
        rows.sort(key=lambda r: r.day)
        return rows


class ExcelDataSource(DataSource):
    """Чтение курсов валют из XLSX-файла (первая строка — заголовки)."""

    def load(self, path: str) -> List[CurrencyRate]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError(
                "Для чтения xlsx установите пакет openpyxl (pip install openpyxl)"
            ) from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        currencies = [str(c) for c in header[1:]]

        rows: List[CurrencyRate] = []
        for raw in rows_iter:
            if raw[0] is None:
                continue
            day = raw[0] if hasattr(raw[0], "year") else _parse_date(str(raw[0]))
            if hasattr(day, "date") and not hasattr(day, "day"):
                day = day.date()
            rates = {cur: float(val) for cur, val in zip(currencies, raw[1:])}
            rows.append(CurrencyRate(day=day, rates=rates))
        rows.sort(key=lambda r: r.day)
        return rows


def _parse_date(text: str):
    text = text.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Не удалось разобрать дату: {text}")
